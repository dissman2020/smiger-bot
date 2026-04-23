import io
import struct
import logging

logger = logging.getLogger(__name__)


def parse_pdf(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def parse_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_doc(content: bytes) -> str:
    """Parse legacy .doc (OLE2/Compound File) format.

    Extracts UTF-16LE encoded text (Chinese + ASCII) from the binary stream.
    Falls back to olefile stream reading if available.
    """
    try:
        import olefile
        if olefile.isOleFile(io.BytesIO(content)):
            return _parse_doc_ole(content)
    except ImportError:
        pass

    return _parse_doc_raw(content)


def _parse_doc_ole(content: bytes) -> str:
    """Extract text from .doc using OLE streams."""
    import olefile

    ole = olefile.OleFileIO(io.BytesIO(content))
    try:
        raw = ole.openstream("WordDocument").read()
    finally:
        ole.close()
    return _extract_utf16_text(raw)


def _parse_doc_raw(content: bytes) -> str:
    """Brute-force extract UTF-16LE text from raw .doc bytes."""
    return _extract_utf16_text(content)


def _extract_utf16_text(raw: bytes) -> str:
    """Scan binary data for UTF-16LE encoded Chinese/ASCII text segments."""
    text_parts: list[str] = []
    i = 0
    while i < len(raw) - 1:
        cc = struct.unpack_from("<H", raw, i)[0]
        if (
            (0x4E00 <= cc <= 0x9FFF)       # CJK Unified Ideographs
            or (0x3000 <= cc <= 0x303F)     # CJK Symbols
            or (0xFF00 <= cc <= 0xFFEF)     # Fullwidth Forms
            or (0x0020 <= cc <= 0x007E)     # Basic ASCII
            or cc in (0x000A, 0x000D, 0x0009)  # newline / tab
        ):
            text_parts.append(chr(cc))
        i += 2

    result = "".join(text_parts)
    lines = result.split("\r")

    cleaned: list[str] = []
    for line in lines:
        line = line.strip()
        if len(line) < 4:
            continue
        cn = sum(1 for c in line if "\u4e00" <= c <= "\u9fff")
        en = sum(1 for c in line if "A" <= c <= "z" or "0" <= c <= "9" or c in " ,.!?:;/+-=()$%\"'")
        total = len(line)
        if total > 0 and (cn + en) / total > 0.5:
            cleaned.append(line)

    return "\n".join(cleaned)


def parse_excel(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"## Sheet: {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def parse_text(content: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "gbk"):
        try:
            return content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="replace")


PARSERS = {
    "application/pdf": parse_pdf,
    "pdf": parse_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": parse_docx,
    "docx": parse_docx,
    "application/msword": parse_doc,
    "doc": parse_doc,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": parse_excel,
    "xlsx": parse_excel,
    "text/plain": parse_text,
    "txt": parse_text,
    "text/markdown": parse_text,
    "md": parse_text,
    "text/csv": parse_text,
    "csv": parse_text,
    "application/json": parse_text,
    "json": parse_text,
}


def parse_file(content: bytes, filename: str, content_type: str | None = None) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    parser = PARSERS.get(content_type or "") or PARSERS.get(ext)
    if parser is None:
        logger.warning("Unsupported file type: %s (%s), treating as plain text", ext, content_type)
        return parse_text(content)

    return parser(content)
